#!/usr/bin/env python3
"""
ExcelデータをSupabase用のSQL INSERT文に変換するスクリプト。
発注管理シートのクライアントは独自にclient+orderを作成し、そこにprogram/sessionを紐づける。

使い方:
  python3 scripts/import-excel.py > scripts/import-data.sql
"""

import openpyxl
import re
from datetime import datetime

BASE_DIR = "/Volumes/1TB_SSD_001/100_MH/MH_AIリスキリング研修_管理"

def esc(val):
    if val is None:
        return "NULL"
    s = str(val).replace("'", "''").strip()
    if not s or s == '-':
        return "NULL"
    return f"'{s}'"

def esc_date(val):
    if val is None:
        return "NULL"
    if isinstance(val, datetime):
        return f"'{val.strftime('%Y-%m-%d')}'"
    s = str(val).strip()
    if not s or s == '-':
        return "NULL"
    m = re.match(r'^(\d{4}-\d{2}-\d{2})', s)
    if m:
        return f"'{m.group(1)}'"
    return "NULL"

def esc_num(val):
    if val is None:
        return "NULL"
    s = str(val).strip()
    m = re.match(r'^(\d+)', s)
    if m:
        return m.group(1)
    try:
        return str(int(float(val)))
    except (ValueError, TypeError):
        return "NULL"

def esc_time(val):
    if val is None:
        return "NULL"
    s = str(val).strip()
    if not s or s == '-':
        return "NULL"
    m = re.match(r'(\d{1,2}):(\d{2})', s)
    if m:
        return f"'{int(m.group(1)):02d}:{m.group(2)}'"
    return "NULL"

def parse_time_range(val):
    if val is None:
        return "NULL", "NULL"
    s = str(val).strip()
    if not s or s == '-':
        return "NULL", "NULL"
    m = re.match(r'(\d{1,2}:\d{2})\s*[-~\u301c\uff5e]\s*(\d{1,2}:\d{2})', s)
    if m:
        return esc_time(m.group(1)), esc_time(m.group(2))
    return "NULL", "NULL"

def map_status(status_text):
    if not status_text:
        return "'1'"
    s = str(status_text).strip()
    if s == '完了': return "'完了'"
    if s == '中止': return "'中止'"
    if s == '失注': return "'失注'"
    if s == '進行中': return "'10'"
    m = re.match(r'^(\d+)\.', s)
    if m: return f"'{m.group(1)}'"
    return "'1'"

def map_training_status(s):
    if not s: return "'受注済'"
    if '完了' in str(s): return "'研修完了'"
    return "'受注済'"

def map_venue_type(venue):
    if not venue: return "'オンライン'"
    s = str(venue).strip()
    if ('オンライン' in s) and ('対面' in s or 'リアル' in s or 'ハイブリ' in s):
        return "'ハイブリッド'"
    if 'オンライン' in s:
        return "'オンライン'"
    if '対面' in s or 'リアル' in s:
        return "'対面'"
    if re.search(r'[都道府県市区町村]', s):
        return "'対面'"
    return "'オンライン'"


print("-- ============================================")
print("-- AIリスキリング研修 初期データインポート")
print("-- ============================================")
print()
print("TRUNCATE training_sessions, training_programs, orders, clients CASCADE;")
print()

# ========================================
# 1. 受注管理シート -> clients + orders
# ========================================
wb1 = openpyxl.load_workbook(f"{BASE_DIR}/AIリスキリング_受注管理シート.xlsx", data_only=True)

# AI進行中受注管理シート
ws_orders = wb1["AI進行中受注管理シート"]
order_client_names = set()

print("-- ========== 受注管理シート ==========")

for row_idx in range(2, ws_orders.max_row + 1):
    client_name = ws_orders.cell(row_idx, 2).value
    if not client_name:
        continue
    client_name = str(client_name).strip()

    product = ws_orders.cell(row_idx, 1).value
    status = ws_orders.cell(row_idx, 4).value
    memo = ws_orders.cell(row_idx, 9).value
    notes = ws_orders.cell(row_idx, 10).value
    col_k = ws_orders.cell(row_idx, 11).value
    col_l = ws_orders.cell(row_idx, 12).value
    schedule = ws_orders.cell(row_idx, 7).value

    pt = str(product).strip() if product else "AIリスキリング"
    if pt not in ("AIリスキリング", "AI開発", "Hero AIVO"):
        pt = "AIリスキリング"

    extra = []
    if col_k: extra.append(str(col_k).strip())
    if col_l: extra.append(str(col_l).strip())
    notes_combined = "; ".join(filter(None, [str(notes).strip() if notes else None] + extra)) if (notes or extra) else None

    if client_name not in order_client_names:
        print(f"INSERT INTO clients (name) VALUES ({esc(client_name)}) ON CONFLICT DO NOTHING;")
        order_client_names.add(client_name)

    print(f"""INSERT INTO orders (client_id, product_type, status, training_schedule_text, memo, notes)
SELECT c.id, {esc(pt)}, {map_status(status)}, {esc(schedule)}, {esc(memo)}, {esc(notes_combined)}
FROM clients c WHERE c.name = {esc(client_name)} LIMIT 1;""")

# 全体受注管理シート
ws_overall = wb1["2025年12月〜全体受注管理"]
for row_idx in range(2, ws_overall.max_row + 1):
    client_name = ws_overall.cell(row_idx, 3).value
    if not client_name:
        continue
    client_name = str(client_name).strip()
    if client_name in order_client_names:
        continue

    product = ws_overall.cell(row_idx, 1).value
    order_date = ws_overall.cell(row_idx, 2).value
    status = ws_overall.cell(row_idx, 4).value
    amount = ws_overall.cell(row_idx, 5).value

    pt = str(product).strip() if product else "AIリスキリング"
    if pt not in ("AIリスキリング", "AI開発", "Hero AIVO"):
        pt = "Hero AIVO" if ("Hero" in pt or "AIVO" in pt) else "AIリスキリング"

    print(f"INSERT INTO clients (name) VALUES ({esc(client_name)}) ON CONFLICT DO NOTHING;")
    order_client_names.add(client_name)

    print(f"""INSERT INTO orders (client_id, product_type, status, order_date, order_amount)
SELECT c.id, {esc(pt)}, {map_status(status)}, {esc_date(order_date)}, {esc_num(amount) if amount else 'NULL'}
FROM clients c WHERE c.name = {esc(client_name)} LIMIT 1;""")

# ========================================
# 2. 発注管理シート -> clients + orders + programs + sessions
#    名前が一致しないものは新規client+orderを作る
# ========================================
wb2 = openpyxl.load_workbook(f"{BASE_DIR}/メタヒーローズ_研修発注管理.xlsx", data_only=True)
ws2 = wb2["シート1"]

print()
print("-- ========== 発注管理シート ==========")

for row_idx in range(3, ws2.max_row + 1):
    company_name = ws2.cell(row_idx, 2).value
    if not company_name:
        continue
    company_name = str(company_name).strip()

    company_url = ws2.cell(row_idx, 3).value
    industry = ws2.cell(row_idx, 4).value
    meeting_notes = ws2.cell(row_idx, 5).value
    pre_survey = ws2.cell(row_idx, 6).value
    status = ws2.cell(row_idx, 7).value
    target = ws2.cell(row_idx, 8).value
    materials = ws2.cell(row_idx, 9).value
    duration = ws2.cell(row_idx, 10).value
    venue = ws2.cell(row_idx, 11).value
    num_part = ws2.cell(row_idx, 12).value
    pre_mtg_person = ws2.cell(row_idx, 13).value
    pre_mtg_date = ws2.cell(row_idx, 14).value
    archive = ws2.cell(row_idx, 34).value
    post_survey = ws2.cell(row_idx, 35).value
    notes = ws2.cell(row_idx, 36).value

    # clientが存在しなければ作成
    if company_name not in order_client_names:
        print(f"INSERT INTO clients (name, url, industry) VALUES ({esc(company_name)}, {esc(company_url)}, {esc(industry)});")
        order_client_names.add(company_name)
        # orderも作成
        print(f"""INSERT INTO orders (client_id, product_type, status)
SELECT c.id, 'AIリスキリング', {map_status(status)}
FROM clients c WHERE c.name = {esc(company_name)} LIMIT 1;""")
    else:
        # URL/industryを更新
        if company_url or industry:
            updates = []
            if company_url: updates.append(f"url = {esc(company_url)}")
            if industry: updates.append(f"industry = {esc(industry)}")
            print(f"UPDATE clients SET {', '.join(updates)} WHERE name = {esc(company_name)};")

    # training_program
    vt = map_venue_type(venue)
    np_val = esc_num(num_part) if num_part else "NULL"
    ts = map_training_status(status)

    print(f"""INSERT INTO training_programs (order_id, status, training_target, duration_text, venue, venue_type, num_participants, pre_meeting_person, pre_meeting_date, meeting_notes_url, pre_survey_url, training_materials_url, archive_url, post_survey_url, notes)
SELECT o.id, {ts}, {esc(target)}, {esc(duration)}, {esc(venue)}, {vt}, {np_val}, {esc(pre_mtg_person)}, {esc_date(pre_mtg_date)}, {esc(meeting_notes)}, {esc(pre_survey)}, {esc(materials)}, {esc(archive)}, {esc(post_survey)}, {esc(notes)}
FROM orders o
JOIN clients c ON o.client_id = c.id
WHERE c.name = {esc(company_name)}
ORDER BY o.created_at DESC LIMIT 1;""")

    # sessions
    for day in range(1, 8):
        col_offset = (day - 1) * 3
        instructor_col = 16 + col_offset
        date_col = 17 + col_offset
        time_col = 18 + col_offset

        if instructor_col > 33:
            break

        instructor = ws2.cell(row_idx, instructor_col).value
        session_date = ws2.cell(row_idx, date_col).value
        session_time = ws2.cell(row_idx, time_col).value

        if not instructor and not session_date:
            continue

        date_val = esc_date(session_date)
        start_t, end_t = parse_time_range(session_time)
        instr = esc(instructor)

        print(f"""INSERT INTO training_sessions (training_program_id, day_number, session_date, start_time, end_time, instructor)
SELECT tp.id, {day}, {date_val}, {start_t}, {end_t}, {instr}
FROM training_programs tp
JOIN orders o ON tp.order_id = o.id
JOIN clients c ON o.client_id = c.id
WHERE c.name = {esc(company_name)}
ORDER BY tp.created_at DESC LIMIT 1;""")

print()
print("-- Done!")
